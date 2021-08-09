import openpyxl

class HandExcel:
    def creat_excel(self):
        openpyxl.Workbook().save('/Users/hudongqi/PycharmProjects/Appium/file/new_account1.xlsx')

    def load_excel(self):
        open_excel = openpyxl.load_workbook('/Users/hudongqi/PycharmProjects/Appium/file/new_account1.xlsx')
        return open_excel

    def get_sheet_data(self, index=None):
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def edit(self,mob):
        wb = self.load_excel()
        sheet = wb.active
        sheet.append([mob])
        wb.save('/Users/hudongqi/PycharmProjects/Appium/file/new_account1.xlsx')

    def get_cell_value(self, row, cols):
        data = self.get_sheet_data().cell(row=row, column=cols).value
        return data

    def row_value(self, row):
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

# HandExcel().creat_excel()
